import requests
import openpyxl
import random
import time
import os
from tkinter import *
from tkinter.ttk import *
from bs4 import BeautifulSoup

#add if PW check if CP check

random.seed()

#stores call signs
callSigns = []
#stores infomation that will be printed into xlsx file
required_info = []
#row that should start pulling and writing information on
rowCount = 2

#xlsx file to be used
os.chdir('C:/Users/A19266/Desktop/Python/Scrapper/')
file_name = "MOU agreements.xlsx"
wb = openpyxl.load_workbook(file_name)
sheets = wb.sheetnames
ws = wb[sheets[0]]
rows = ws.max_row

#parses the xlsx excel file and adds all the FCC callsign urls in colomn 2 to the callSigns list
def parser():
    x=1
    while x <= rows:
        if ws.cell(row=x, column=2).hyperlink != None:
            callSigns.append(ws.cell(row=x, column=2).hyperlink.target)
        x+=1

#scrapes and formats the information into required_info
def contactScraper(url):
    response = requests.get(url, verify=False)
    soup = BeautifulSoup(response.content, "html.parser")
    elements = soup.find_all(class_="cell-pri-light")
    items = [e.find_next(class_="cell-pri-light") for e in elements]
    items.pop()

    print (f"Scraping: {url}")
    required_info.clear()
    item_list = []

    #Formating and storing scrapped data into required_info
    x = 0
    y = 99
    z = 99
    for item in items:
    
        item_text = item.get_text()
        item_text = item_text.strip()

        #filters blanks
        if item_text == '':
            continue
        else:
            item_list.append(item_text)
        
        #grabs expiration date
        if x == 4:
            expiration_date = item_list[x]
            required_info.append(expiration_date)

        #checks for text and stops appending to required_info
        if item_text == 'Mobile' or item_text == 'Fixed' or item_text == 'Fixed, Mobile' or item_text == 'Private Comm':
            z = x

        #tells program when to append to required_info
        if x > y and x < z:
            required_info.append(item_text)
            x += 1
            if x <= z:
                x += 1
                continue
    
        #checks for text and starts appending to required_info
        if item_text == 'Governmental Entity' or item_text == 'Corporation' or item_text == 'Individual':
            y = x

        x += 1
    
    #removes unnecessary items from required_info
    required_info.pop()
    #print(required_info)
    #print("list length:", len(required_info))
    #print(item_list) 

#call contact scraper and writes to the xlsx file
def scraper():
    global rowCount
    
    for x in callSigns:
        if x == 'NA':
            rowCount += 1
            continue
        else:
            print('')
            contactScraper(x)
            writer()
            rowCount += 1
            wait_timer = random.randint(5,10)
            print("Sleep for: ", wait_timer)
            time.sleep(wait_timer)
        
#writes to xlsx file
def writer():
    colCount=3
    infoCount=0
    for element in required_info:
        ws.cell(row=rowCount, column=colCount).value = required_info[infoCount]
        colCount+=1
        infoCount+=1

def main():
    
    print("current working directory: ", os.getcwd())
    
    parser()
    scraper()
    wb.save("MOU agreements.xlsx")

if __name__ == "__main__":
    main()