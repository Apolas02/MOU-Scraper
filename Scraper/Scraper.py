import requests
import openpyxl
import random
import os
import sys
from datetime import datetime
from bs4 import BeautifulSoup

random.seed()
#list of are callSigns
areaCallSigns = []
#stores infomation that will be printed into xlsx file
required_info = []
#row that should start pulling and writing information on
rowCount = 2

#preps the xlsx file to be used if xlsx file is missing or named incorrect, will throw and exception
os.chdir(sys.path[0])
file_name = "MOU agreements.xlsx"
try:
    wb = openpyxl.load_workbook(file_name)
except:
    print("MOU agreements.xlsx not found")
    input("press enter key to exit")
    exit()
sheets = wb.sheetnames
ws = wb[sheets[0]]


#parses the xlsx excel file and adds all the FCC callsign urls in colomn 2 to the callSigns list
def parser():
    callSigns = []
    rows = ws.max_row
    x=2
    while x <= rows:
        callSignLink = ws.cell(row=x, column=2).hyperlink
        if ws.cell(row=x, column=2).hyperlink != None:
            callSigns.append(callSignLink.target)
        else:
            callSigns.append("NA")
        x+=1
    areaCallSigns.append(callSigns)

#scrapes and formats the information into required_info
def contactScraper(url):
    print (f"Scraping: {url}")
    response = requests.get(url, verify=False)
    soup = BeautifulSoup(response.content, "html.parser")
    elements = soup.find_all(class_="cell-pri-light")
    items = [e.find_next(class_="cell-pri-light") for e in elements]
    entity_type = ['Governmental Entity', 'Corporation', 'Individual', 'Limited Liability Company']
    svc_type = ['Mobile', 'Fixed', 'Fixed, Mobile', 'Private Comm', 'FCL']

    if len(items) == 0:
        return
    else:    
        items.pop()
        required_info.clear()
        item_list = []

        #Formating and storing scrapped data into required_info
        x = 0
        radio_svc_type_check = False #99
        entity_type_check = False #99
        for item in items:
        
            item_text = item.get_text()
            item_text = item_text.strip()

            #filters blanks
            if item_text == '':
                continue
            else:
                item_list.append(item_text)
            
            if x == 0:
                radio_service = item_text[0:2]

            #grabs expiration date
            if x == 4 and radio_service != 'CP':
                expiration_date = item_list[x]
                required_info.append(expiration_date)

            if x == 8 and radio_service == 'CP':
                expiration_date = item_list[x]
                required_info.append(expiration_date)

            #checks for text and stops appending to required_info
            if item_text in svc_type:
                radio_svc_type_check = True

            #tells program when to append to required_info
            if radio_svc_type_check == False and entity_type_check == True:
                required_info.append(item_text)
                x += 1
        
            #checks for text and starts appending to required_info
            if item_text in entity_type:
                entity_type_check = True
            x += 1

            if radio_svc_type_check == True and entity_type_check == True:
                break
        
        #formats then removes unnecessary items from required_info
        formater()
        if radio_service != 'MC':
            required_info.pop()

        if radio_service == 'CF':
            required_info.pop()

        if len(required_info) == 7:
            required_info[5] = required_info[6]
            required_info.pop()
        
        if len(required_info) == 6:
           if required_info[5] == 'Yes' or required_info[5] == 'No':
                required_info.pop()
        writer()

#call contactScraper and writes to the xlsx file
def scraper(currentCallSigns):
    global rowCount
    for x in currentCallSigns:
        if x == 'NA':
            rowCount += 1
            continue
        else:
            print('')
            contactScraper(x)
            rowCount += 1
            #wait_timer = random.randint(5,10)
            #print("Sleep for: ", wait_timer)
            #time.sleep(wait_timer)
    rowCount = 2
    
#the main call of the program, it iterates through the sheets of the xlsx file and calls the parse and scraper functions
def sheet_changer():
    global ws
    y = 0
    for x in sheets:
        print('')
        ws = wb[sheets[y]]
        parser()
        print(x, "division has", len(areaCallSigns[y]), "callsigns")
        scraper(areaCallSigns[y])
        y+=1

#writes to xlsx file
def writer():
    colCount=3
    infoCount=0
    for element in required_info:
        ws.cell(row=rowCount, column=colCount).value = required_info[infoCount]
        ws.cell(row=rowCount, column=colCount).alignment = openpyxl.styles.Alignment(wrap_text=True)
        colCount+=1
        infoCount+=1

#formats the text in required_info
def formater():
    required_info[1] = required_info[1].splitlines()
    required_info[3] = required_info[3].splitlines()
    required_info[0] = datetime.strptime(required_info[0], '%m/%d/%Y').date()

    ###LICENSEE ADDRESS BUILDING###
    lic_po_box = ''
    lic_address = ''
    lic_street_number = ''
    lic_city_state = ''
    lic_zip_code = ''
    lic_extra_info = ''

    #No PO box, no ATTN
    if len(required_info[1]) == 3:
        lic_street_number = required_info[1][0]
        lic_city_state = required_info[1][1]
        lic_zip_code = required_info[1][2]
        lic_address = lic_street_number + ', ' + lic_city_state + ', ' + lic_zip_code

    #PO box, no ATTN
    elif len(required_info[1]) == 4:
        lic_po_box = required_info[1][0]
        lic_street_number = required_info[1][1]
        lic_city_state = required_info[1][2]
        lic_zip_code = required_info[1][3]
        lic_address = lic_po_box + ', ' + lic_street_number + ', ' + lic_city_state + ', ' + lic_zip_code
    
    #No PO box, w/ ATTN
    elif len(required_info[1]) == 6:
        lic_street_number = required_info[1][1]
        lic_city_state = required_info[1][2]
        lic_zip_code = required_info[1][3]
        lic_extra_info = required_info[1][5]
        lic_address = lic_street_number + ', ' + lic_city_state + ', ' + lic_zip_code + '\n' + lic_extra_info

    #PO box, w/ ATTN
    elif len(required_info[1]) == 7:
        lic_po_box = required_info[1][1]
        lic_street_number = required_info[1][2]
        lic_city_state = required_info[1][3]
        lic_zip_code = required_info[1][4]
        lic_extra_info = required_info[1][6]
        lic_address = lic_po_box + ', ' + lic_street_number + ', ' + lic_city_state + ', ' + lic_zip_code + '\n' + lic_extra_info
    required_info[1] = lic_address


    ###CONTACT NAME BUILDING###
    con_top = []
    entity = ''
    first_name = ''
    middle_initial = ''
    last_name = ''
    prefix = ''
    con_name = ''

    ###CONTACT ADDRESS BUILDING###
    con_bottom = []
    con_po_box = ''
    con_street_number = ''
    con_city_state = ''
    con_zip_code = ''
    con_extra_info = ''
    con_address = ''
    
    #split contact into top half and bottom half
    contact_switcher = False
    for x in required_info[3]:
        if len(x) == 0:
            contact_switcher = True
            continue
        
        if contact_switcher == False:
            con_top.append(x)
        else:
            con_bottom.append(x)
    
    if len(con_top) == 1:
        entity = con_top[0]
        con_name = entity
        

    elif len(con_top) == 2:
        first_name = con_top[0]
        last_name = con_top[1]
        con_name = first_name + ' ' + last_name
        
    elif len(con_top) == 3:
        entity = con_top[0]
        first_name = con_top[1]
        last_name = con_top[2]
        con_name = entity + '\n' + first_name + ' ' + last_name
    
    elif len(con_top) == 4:
        entity = con_top[0]
        first_name = con_top[1]
        middle_initial = con_top[2]
        last_name = con_top[3]
        con_name = entity + '\n' + first_name + ' ' + middle_initial + last_name

    elif len(con_top) == 5:
        entity = con_top[0]
        first_name = con_top[1]
        middle_initial = con_top[2]
        last_name = con_top[3] 
        prefix = con_top[4]
        con_name = entity + '\n' + prefix + first_name + ' ' + middle_initial + last_name
    

    if len(con_bottom) == 3:
        con_street_number = con_bottom[0]
        con_city_state = con_bottom[1]
        con_zip_code = con_bottom[2]
        con_address = con_street_number + ', ' + con_city_state + ', ' + con_zip_code
    
    elif len(con_bottom) == 4:
        con_street_number = con_bottom[0]
        con_city_state = con_bottom[1]
        con_zip_code = con_bottom[2]
        con_extra_info = con_bottom[3]
        con_address = con_street_number + ', ' + con_city_state + ', ' + con_zip_code + '\n' + con_extra_info
    
    elif len(con_bottom) == 5:
        con_po_box = con_bottom[0]
        con_street_number = con_bottom[1]
        con_city_state = con_bottom[2]
        con_zip_code = con_bottom[3]
        con_extra_info = con_bottom[4]
        con_address = con_po_box + ', ' + con_street_number + ', ' + con_city_state + ', ' + con_zip_code + '\n' + con_extra_info
    required_info[3] = con_name + '\n' + con_address


def main():
    #print("current working directory: ", os.getcwd())
    sheet_changer()
    wb.save('MOU agreements.xlsx')
    print("\nScraping complete, MOU agreements.xlsx has been updated.")
    input("press enter key to exit")
    

if __name__ == "__main__":
    main()